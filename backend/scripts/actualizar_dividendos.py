"""
actualizar_dividendos.py
========================
Actualiza el Excel de dividendos con los datos del alpaca_data.json.

Uso:
    python actualizar_dividendos.py alpaca_data.json dividendos_calendario.xlsx

Qué actualiza:
    - Cantidad de acciones (de holdings)
    - Precio de mercado (de holdings)
"""

import json
import sys
import os
from openpyxl import load_workbook

def main():
    if len(sys.argv) < 3:
        print("Uso: python actualizar_dividendos.py <alpaca_data.json> <dividendos.xlsx>")
        sys.exit(1)

    json_path = sys.argv[1]
    xlsx_path = sys.argv[2]

    with open(json_path, 'r') as f:
        data = json.load(f)

    holdings = data.get('holdings', {})
    period = data.get('header', {}).get('period', '?')

    wb = load_workbook(xlsx_path)
    ws = wb['Resumen']

    # Encontrar la fila de headers para ubicar columnas
    hdr_row = None
    for row in range(1, 10):
        val = ws.cell(row=row, column=1).value
        if val and 'Ticker' in str(val):
            hdr_row = row
            break

    if not hdr_row:
        print("ERROR: No se encontró la fila de headers en hoja Resumen")
        sys.exit(1)

    # Mapear tickers a filas del Excel
    ticker_rows = {}
    r = hdr_row + 1
    while True:
        tk = ws.cell(row=r, column=1).value
        if tk is None or tk == '' or tk == 'TOTAL':
            break
        ticker_rows[str(tk).strip().upper()] = r
        r += 1

    updated = []
    not_found = []

    for ticker, info in holdings.items():
        tk = ticker.strip().upper()
        if tk in ticker_rows:
            row = ticker_rows[tk]
            old_qty = ws.cell(row=row, column=3).value
            old_price = ws.cell(row=row, column=4).value

            new_qty = info.get('qty', old_qty)
            new_price = info.get('market_price', old_price)

            ws.cell(row=row, column=3).value = new_qty
            ws.cell(row=row, column=4).value = new_price

            updated.append(f"  {tk:6s}  qty: {old_qty:.4f} -> {new_qty:.6f}  |  price: ${old_price:.2f} -> ${new_price:.2f}")
        else:
            not_found.append(tk)

    wb.save(xlsx_path)

    print(f"Periodo: {period}")
    print(f"Actualizados: {len(updated)} tickers")
    for line in updated:
        print(line)

    if not_found:
        print(f"\nNo encontrados en el Excel (agregar manualmente): {', '.join(not_found)}")

    print(f"\nArchivo guardado: {xlsx_path}")
    print("TIP: Recalcular fórmulas abriendo en Excel/LibreOffice o con recalc.py")

if __name__ == '__main__':
    main()